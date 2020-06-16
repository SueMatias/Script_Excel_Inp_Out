from openpyxl import load_workbook
from openpyxl.styles import PatternFill

path = '/home/suellen/Documentos/IC/SrCeO.xlsx'
arquivo_excel = load_workbook(path)

column = 2
for Sr in range(1, 8):
    for Ce in range(1, 2):
        for O in range(1, 30):
            volume = open(f"/home/suellen/Documentos/IC/SCO_WC1LYP/{Sr}{Ce}{O}/SCO_WC1LYP_{Sr}{Ce}{O}_VOLUME_TROCA.txt", "r")
            parametro_A = open(f"/home/suellen/Documentos/IC/SCO_WC1LYP/{Sr}{Ce}{O}/SCO_WC1LYP_{Sr}{Ce}{O}_PARAMETRO_TROCA.txt", "r")
            parametro_B = open(f"/home/suellen/Documentos/IC/SCO_WC1LYP/{Sr}{Ce}{O}/SCO_WC1LYP_{Sr}{Ce}{O}_PARAMETRO_TROCA_B.txt", "r")
            parametro_C = open(f"/home/suellen/Documentos/IC/SCO_WC1LYP/{Sr}{Ce}{O}/SCO_WC1LYP_{Sr}{Ce}{O}_PARAMETRO_TROCA_C.txt", "r")
            gap = open(f"/home/suellen/Documentos/IC/SCO_WC1LYP/{Sr}{Ce}{O}/SCO_WC1LYP_{Sr}{Ce}{O}_GAP_TROCA.txt", "r")
            valor_gap = []
            valor_volume = []
            valor_parametro_A = []
            valor_parametro_B = []
            valor_parametro_C = []
            for linha in volume:
                linha = linha.strip()
                valor_volume.append(linha)
            for linha in parametro_A:
                linha = linha.strip()
                valor_parametro_A.append(linha)
            for linha in parametro_B:
                linha = linha.strip()
                valor_parametro_B.append(linha)
            for linha in parametro_C:
                linha = linha.strip()
                valor_parametro_C.append(linha)
            for linha in gap:
                linha = linha.strip()
                valor_gap.append(linha)

            if len(valor_gap) != 1:
                WC1LYP = arquivo_excel.active
                redFill = PatternFill(start_color="FFFF0000",
                                      end_color="FFFF0000",
                                      fill_type="solid")
                for row in range(2, 11):
                    WC1LYP.cell(row=row, column=column).fill = redFill
            else:
                print(f'Combinação = {Sr}{Ce}{O}')

                WC1LYP = arquivo_excel.active
                WC1LYP.cell(row=2, column=column).value = valor_volume[0]
                print(WC1LYP.cell(row=2, column=column).value)

                WC1LYP.cell(row=4, column=column).value = valor_parametro_A[0]
                print(WC1LYP.cell(row=4, column=column).value)

                WC1LYP.cell(row=6, column=column).value = valor_parametro_B[0]
                print(WC1LYP.cell(row=6, column=column).value)

                WC1LYP.cell(row=8, column=column).value = valor_parametro_C[0]
                print(WC1LYP.cell(row=8, column=column).value)

                WC1LYP.cell(row=12, column=column).value = valor_gap[0]
                print(WC1LYP.cell(row=12, column=column).value)

            column += 1

arquivo_excel.save("/home/suellen/Documentos/IC/SrCeO.xlsx")
