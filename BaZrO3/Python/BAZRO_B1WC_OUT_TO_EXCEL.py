from openpyxl import load_workbook

path = '/home/suellen/Documentos/IC/BaZrO.xlsx'
arquivo_excel = load_workbook(path)

column = 2
for Ba in range(1, 8):
    for Z in range(1, 7):
        for O in range(1, 30):
            volume = open(f"/home/suellen/Documentos/IC/BAZRO_B1WC/{Ba}{Z}{O}/BZO_B1WC_{Ba}{Z}{O}_VOLUME_TROCA.txt", "r")
            parametro = open(f"/home/suellen/Documentos/IC/BAZRO_B1WC/{Ba}{Z}{O}/BZO_B1WC_{Ba}{Z}{O}_PARAMETRO_TROCA.txt", "r")
            densidade = open(f"/home/suellen/Documentos/IC/BAZRO_B1WC/{Ba}{Z}{O}/BZO_B1WC_{Ba}{Z}{O}_DENSITY_TROCA.txt", "r")
            gap = open(f"/home/suellen/Documentos/IC/BAZRO_B1WC/{Ba}{Z}{O}/BZO_B1WC_{Ba}{Z}{O}_GAP_TROCA.txt", "r")
            valor_volume = []
            valor_parametro = []
            valor_densidade = []
            valor_gap = []
            for linha in volume:
                linha = linha.strip()
                valor_volume.append(linha)
            for linha in parametro:
                linha = linha.strip()
                valor_parametro.append(linha)
            for linha in densidade:
                linha = linha.strip()
                valor_densidade.append(linha)
            for linha in gap:
                linha = linha.strip()
                valor_gap.append(linha)

            if len(valor_gap) != 1:
                pass
            else:
                print(f'Combinação = {Ba}{Z}{O}')

                B1WC = arquivo_excel.active
                if Ba <= 4:
                    B1WC.cell(row=2, column=column).value = valor_volume[0]
                    print(B1WC.cell(row=2, column=column).value)

                    B1WC.cell(row=4, column=column).value = valor_parametro[0]
                    print(B1WC.cell(row=4, column=column).value)

                    B1WC.cell(row=6, column=column).value = valor_gap[0]
                    print(B1WC.cell(row=6, column=column).value)
                    
                    B1WC.cell(row=8, column=column).value = valor_densidade[0]
                    print(B1WC.cell(row=8, column=column).value)
                else:
                    B1WC.cell(row=14, column=column).value = valor_volume[0]
                    print(B1WC.cell(row=2, column=column).value)

                    B1WC.cell(row=16, column=column).value = valor_parametro[0]
                    print(B1WC.cell(row=4, column=column).value)

                    B1WC.cell(row=18, column=column).value = valor_gap[0]
                    print(B1WC.cell(row=6, column=column).value)
                    
                    B1WC.cell(row=20, column=column).value = valor_densidade[0]
                    print(B1WC.cell(row=8, column=column).value)

            column += 1

arquivo_excel.save("/home/suellen/Documentos/IC/BaZrO.xlsx")
